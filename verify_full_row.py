#!/usr/bin/env python3
"""
Verify that full-row highlighting is working
"""

from openpyxl import load_workbook

def verify_full_row_highlighting(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    
    highlighted_rows = {}
    max_colored_col = 0
    
    for row_idx in range(1, min(ws.max_row + 1, 50)):  # Check first 50 rows
        row_colors = set()
        colored_columns = 0
        
        for col_idx in range(1, 20):  # Check first 20 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if (cell.fill and 
                hasattr(cell.fill, 'fgColor') and 
                cell.fill.fgColor and 
                hasattr(cell.fill.fgColor, 'rgb') and 
                cell.fill.fgColor.rgb and 
                cell.fill.patternType != 'none' and
                cell.fill.fgColor.rgb not in ['00000000', 'FFFFFFFF']):
                
                row_colors.add(cell.fill.fgColor.rgb)
                colored_columns += 1
                max_colored_col = max(max_colored_col, col_idx)
        
        if row_colors:
            highlighted_rows[row_idx] = {
                'colors': list(row_colors),
                'colored_columns': colored_columns,
                'data_value': str(ws.cell(row=row_idx, column=1).value)[:30] if ws.cell(row=row_idx, column=1).value else ''
            }
    
    return highlighted_rows, max_colored_col

# Test the new full-row highlighting
file_path = "/Users/gr4yf1r3/Library/CloudStorage/OneDrive-Nuance/audioMover/_migration/walgreens_excelPlayground/ReDooV2/en_transcriptions_locationprompt_Tuned_9.15.2025_For_ScriptSplits_part1-4/sample_data/gui_test_combined.xlsx"

print("Checking full-row highlighting...")
highlighted_rows, max_col = verify_full_row_highlighting(file_path)

print(f"\nFound {len(highlighted_rows)} highlighted rows")
print(f"Maximum colored column: {max_col}")
print("\nFirst 10 highlighted rows:")

for i, (row_num, info) in enumerate(list(highlighted_rows.items())[:10]):
    colors_str = ', '.join(info['colors'])
    print(f"  Row {row_num}: {info['colored_columns']} colored columns, Colors: {colors_str}")
    print(f"    Data: {info['data_value']}")

# Check if we have good row coverage
rows_with_wide_coverage = sum(1 for info in highlighted_rows.values() if info['colored_columns'] >= 10)
print(f"\nRows with wide highlighting (10+ columns): {rows_with_wide_coverage}")

if rows_with_wide_coverage > 0:
    print("✅ Full-row highlighting is working!")
else:
    print("❌ Full-row highlighting needs improvement")