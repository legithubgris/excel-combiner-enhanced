#!/usr/bin/env python3
"""
Excel File Combiner GUI Application

A cross-platform GUI application for combining multiple Excel files into a single file.
Built with tkinter for maximum compatibility across macOS and Windows.
"""

import os
import sys
import glob
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
from pathlib import Path
import threading
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import copy

class ExcelCombinerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel File Combiner")
        self.root.geometry("800x600")
        self.root.minsize(600, 400)
        
        # Variables
        self.folder_path = tk.StringVar()
        self.output_filename = tk.StringVar(value="combined_excel_files.xlsx")
        self.is_processing = False
        
        # Set up the GUI
        self.setup_gui()
        
        # Center the window
        self.center_window()
    
    def center_window(self):
        """Center the window on the screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def setup_gui(self):
        """Set up the GUI components"""
        
        # Main frame with padding
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(4, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="Excel File Combiner", 
                               font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Folder selection
        ttk.Label(main_frame, text="Source Folder:").grid(row=1, column=0, sticky=tk.W, pady=5)
        
        folder_frame = ttk.Frame(main_frame)
        folder_frame.grid(row=1, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        folder_frame.columnconfigure(0, weight=1)
        
        self.folder_entry = ttk.Entry(folder_frame, textvariable=self.folder_path, width=50)
        self.folder_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 5))
        
        self.browse_button = ttk.Button(folder_frame, text="Browse", command=self.browse_folder)
        self.browse_button.grid(row=0, column=1)
        
        # Output filename
        ttk.Label(main_frame, text="Output File:").grid(row=2, column=0, sticky=tk.W, pady=5)
        
        output_frame = ttk.Frame(main_frame)
        output_frame.grid(row=2, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        output_frame.columnconfigure(0, weight=1)
        
        self.output_entry = ttk.Entry(output_frame, textvariable=self.output_filename, width=50)
        self.output_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 5))
        
        # Buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=3, pady=20)
        
        self.combine_button = ttk.Button(button_frame, text="Combine Excel Files", 
                                        command=self.start_combine_process, style="Accent.TButton")
        self.combine_button.pack(side=tk.LEFT, padx=5)
        
        self.clear_button = ttk.Button(button_frame, text="Clear Log", command=self.clear_log)
        self.clear_button.pack(side=tk.LEFT, padx=5)
        
        # Progress bar
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(50, 10))
        
        # Log text area
        log_frame = ttk.LabelFrame(main_frame, text="Processing Log", padding="5")
        log_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, wrap=tk.WORD)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # Initial log message
        self.log_message("Excel File Combiner started. Select a folder containing Excel files to begin.")
    
    def browse_folder(self):
        """Browse for folder containing Excel files"""
        folder = filedialog.askdirectory(title="Select folder containing Excel files")
        if folder:
            self.folder_path.set(folder)
            self.log_message(f"Selected folder: {folder}")
            
            # Check for Excel files in the selected folder
            excel_files = self.get_excel_files(folder)
            if excel_files:
                self.log_message(f"Found {len(excel_files)} Excel files:")
                for file in excel_files:
                    self.log_message(f"  - {os.path.basename(file)}")
            else:
                self.log_message("No Excel files found in selected folder.")
    
    def get_excel_files(self, folder_path, exclude_files=None):
        """Get all Excel files from the specified folder."""
        if exclude_files is None:
            exclude_files = ['combined_excel_files.xlsx', 'test_combined.xlsx', 
                           'updated_combined.xlsx', 'final_combined.xlsx', 
                           'final_updated_combined.xlsx', 'clean_test.xlsx']
        
        # Add current output filename to exclusions
        output_file = self.output_filename.get()
        if output_file and output_file not in exclude_files:
            exclude_files.append(output_file)
        
        # Look for both .xlsx and .xls files
        xlsx_pattern = os.path.join(folder_path, "*.xlsx")
        xls_pattern = os.path.join(folder_path, "*.xls")
        
        excel_files = glob.glob(xlsx_pattern) + glob.glob(xls_pattern)
        
        # Filter out excluded files
        filtered_files = []
        for file_path in excel_files:
            filename = os.path.basename(file_path)
            if filename not in exclude_files:
                filtered_files.append(file_path)
        
        # Sort files to ensure consistent order
        filtered_files.sort()
        
        return filtered_files
    
    def read_excel_data_with_formatting(self, file_path):
        """Read Excel file and return data with formatting information."""
        try:
            # Read the Excel file with pandas for data
            df = pd.read_excel(file_path)
            
            # Load with openpyxl to get formatting
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Get the filename without extension for the source column
            filename = os.path.basename(file_path)
            
            # Capture row formatting information
            row_formats = {}
            
            for row_idx in range(1, ws.max_row + 1):
                row_format = {}
                has_formatting = False
                
                # Check each cell in the row for formatting
                for col_idx in range(1, min(ws.max_column + 1, len(df.columns) + 1)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    cell_format = {}
                    
                    # Check for fill (background color)
                    if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
                        if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                            if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                                # Store the RGB value as string
                                rgb_val = cell.fill.fgColor.rgb
                                if hasattr(rgb_val, 'rgb'):
                                    rgb_val = rgb_val.rgb
                                cell_format['fill_color'] = str(rgb_val)
                                has_formatting = True
                    
                    # Check for font formatting
                    if cell.font:
                        font_info = {}
                        if cell.font.bold:
                            font_info['bold'] = True
                            has_formatting = True
                        if cell.font.italic:
                            font_info['italic'] = True
                            has_formatting = True
                        # Skip font color for now to avoid errors
                        # if cell.font.color and hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                        #     font_info['color'] = str(cell.font.color.rgb)
                        #     has_formatting = True
                        
                        if font_info:
                            cell_format['font'] = font_info
                    
                    if cell_format:
                        row_format[col_idx] = cell_format
                
                if has_formatting:
                    row_formats[row_idx] = row_format
            
            return df, filename, row_formats
            
        except Exception as e:
            self.log_message(f"Error reading file {file_path}: {str(e)}")
            return None, None, None
    
    def read_excel_data(self, file_path):
        """Read Excel file and return data from columns A through C (legacy method)."""
        result = self.read_excel_data_with_formatting(file_path)
        if result[0] is not None:
            return result[0], result[1]
        return None, None
    
    def combine_excel_files(self):
        """Combine multiple Excel files into one with preserved formatting."""
        folder_path = self.folder_path.get()
        output_filename = self.output_filename.get()
        
        if not folder_path:
            messagebox.showerror("Error", "Please select a source folder.")
            return False
        
        if not os.path.exists(folder_path):
            messagebox.showerror("Error", f"Folder does not exist: {folder_path}")
            return False
        
        if not output_filename:
            messagebox.showerror("Error", "Please specify an output filename.")
            return False
        
        # Get all Excel files in the folder, excluding output files
        excel_files = self.get_excel_files(folder_path)
        
        if not excel_files:
            self.log_message(f"No Excel files found in folder: {folder_path}")
            messagebox.showwarning("Warning", "No Excel files found in the selected folder.")
            return False
        
        self.log_message(f"Found {len(excel_files)} Excel files to combine:")
        for file in excel_files:
            self.log_message(f"  - {os.path.basename(file)}")
        
        try:
            # Create a new workbook for output
            output_wb = Workbook()
            output_ws = output_wb.active
            output_ws.title = "Combined_Data"
            
            # Initialize variables for combining data
            combined_data = []
            all_formatting = []
            current_row = 1
            header_added = False
            
            for file_index, file_path in enumerate(excel_files):
                self.log_message(f"Processing: {os.path.basename(file_path)}")
                
                df, source_filename, row_formats = self.read_excel_data_with_formatting(file_path)
                
                if df is None:
                    continue
                    
                # Skip empty files
                if df.empty:
                    self.log_message(f"  Skipping empty file: {source_filename}")
                    continue
                
                # Ensure we have standard column names
                original_columns = list(df.columns)
                if len(df.columns) >= 3:
                    # Use first 3 columns and add source column
                    df_subset = df.iloc[:, :3].copy()
                    df_subset.columns = ['Filename', 'Transcription', 'Status']
                    df_subset['Source_File'] = ''
                else:
                    self.log_message(f"  Warning: File {source_filename} has fewer than 3 columns. Skipping.")
                    continue
                
                # Add source filename to the first data row of this file
                if len(df_subset) > 0:
                    if not header_added:
                        # First file: include header
                        df_subset.iloc[0, df_subset.columns.get_loc('Source_File')] = source_filename
                        start_data_row = 1
                        header_added = True
                    else:
                        # Subsequent files: skip header, add source to first data row
                        if len(df_subset) > 1:
                            df_subset = df_subset.iloc[1:].copy()  # Skip header row
                            if len(df_subset) > 0:
                                df_subset.iloc[0, df_subset.columns.get_loc('Source_File')] = source_filename
                        start_data_row = 2  # Start from row 2 in original file (skip header)
                
                # Store data and formatting info
                combined_data.append(df_subset)
                
                # Adjust row formatting indices for the combined file
                if row_formats and len(df_subset) > 0:
                    adjusted_formats = {}
                    for orig_row, format_info in row_formats.items():
                        # Calculate new row position in combined file
                        if file_index == 0:
                            # First file
                            new_row = current_row + (orig_row - 1)
                        else:
                            # Subsequent files: adjust for skipped header
                            if orig_row > 1:  # Skip header row
                                new_row = current_row + (orig_row - 2)
                            else:
                                continue  # Skip header formatting
                        
                        adjusted_formats[new_row] = format_info
                    
                    all_formatting.append(adjusted_formats)
                
                current_row += len(df_subset)
                self.log_message(f"  Added {len(df_subset)} rows with formatting")
            
            if not combined_data:
                self.log_message("No data to combine!")
                messagebox.showwarning("Warning", "No data found to combine.")
                return False
            
            # Combine all DataFrames
            final_df = pd.concat(combined_data, ignore_index=True)
            
            # Write data to the workbook
            for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    output_ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Apply formatting with full-row highlighting
            formatted_rows_set = set()  # Track which rows need full-row formatting
            row_colors = {}  # Store the primary color for each row
            
            # First pass: identify rows that need formatting and determine their colors
            for format_dict in all_formatting:
                for row_num, row_format in format_dict.items():
                    if row_num <= output_ws.max_row:
                        # Check if this row has any fill colors
                        for col_num, cell_format in row_format.items():
                            if 'fill_color' in cell_format:
                                try:
                                    fill_color = cell_format['fill_color']
                                    # Ensure it's a valid hex color
                                    if len(fill_color) == 8 and fill_color.startswith('FF'):
                                        fill_color = fill_color[2:]  # Remove alpha channel
                                    elif len(fill_color) != 6:
                                        continue  # Skip invalid colors
                                    
                                    # Skip default/black colors
                                    if fill_color not in ['000000', 'FFFFFF']:
                                        formatted_rows_set.add(row_num)
                                        row_colors[row_num] = fill_color
                                        break  # Use first valid color found for the row
                                except Exception:
                                    continue
            
            # Second pass: apply formatting to entire rows
            max_column = max(output_ws.max_column, 15)  # Ensure we color at least 15 columns for visual effect
            
            # Apply full-row background colors first
            for row_num in formatted_rows_set:
                fill_color = row_colors[row_num]
                try:
                    pattern_fill = PatternFill(start_color=fill_color, 
                                             end_color=fill_color, 
                                             fill_type='solid')
                    
                    # Apply background color to entire row (extend well beyond data columns)
                    for col_idx in range(1, max_column + 10):  # Extra columns for visual effect
                        cell = output_ws.cell(row=row_num, column=col_idx)
                        cell.fill = pattern_fill
                        
                except Exception as e:
                    self.log_message(f"    Warning: Could not apply full-row color {fill_color} to row {row_num}: {e}")
            
            # Third pass: apply font formatting to original cells
            for format_dict in all_formatting:
                for row_num, row_format in format_dict.items():
                    if row_num <= output_ws.max_row:
                        for col_num, cell_format in row_format.items():
                            if col_num <= output_ws.max_column:
                                cell = output_ws.cell(row=row_num, column=col_num)
                                
                                # Apply font formatting (simplified - only bold/italic)
                                if 'font' in cell_format:
                                    try:
                                        font_info = cell_format['font']
                                        current_font = cell.font
                                        
                                        new_font = Font(
                                            name=current_font.name or 'Calibri',
                                            size=current_font.size or 11,
                                            bold=font_info.get('bold', current_font.bold),
                                            italic=font_info.get('italic', current_font.italic)
                                            # Skip color for now due to complexity
                                        )
                                        cell.font = new_font
                                    except Exception as e:
                                        self.log_message(f"    Warning: Could not apply font formatting: {e}")
            
            formatted_rows = formatted_rows_set  # For the count at the end
            
            # Create output file path and save
            output_path = os.path.join(folder_path, output_filename)
            output_wb.save(output_path)
            
            self.log_message(f"Successfully combined {len(excel_files)} files with preserved formatting!")
            self.log_message(f"Output saved to: {output_path}")
            self.log_message(f"Total rows in combined file: {len(final_df)}")
            self.log_message(f"Columns: {list(final_df.columns)}")
            
            # Count formatted rows (now refers to full-row formatting)
            total_formatted_rows = len(formatted_rows)
            self.log_message(f"Applied full-row highlighting to {total_formatted_rows} row(s)")
            
            messagebox.showinfo("Success", 
                              f"Successfully combined {len(excel_files)} files with full-row formatting!\n"
                              f"Output saved to: {output_filename}\n"
                              f"Total rows: {len(final_df)}\n"
                              f"Highlighted rows: {total_formatted_rows}")
            
            return True
            
        except Exception as e:
            error_msg = f"Error combining files: {str(e)}"
            self.log_message(error_msg)
            messagebox.showerror("Error", error_msg)
            return False
    
    def start_combine_process(self):
        """Start the combination process in a separate thread"""
        if self.is_processing:
            return
        
        self.is_processing = True
        self.combine_button.config(state='disabled')
        self.progress.start(10)
        self.status_var.set("Processing...")
        
        # Run in separate thread to prevent GUI freezing
        thread = threading.Thread(target=self.combine_process_thread)
        thread.daemon = True
        thread.start()
    
    def combine_process_thread(self):
        """Thread function for combining files"""
        try:
            success = self.combine_excel_files()
            
            # Update GUI in main thread
            self.root.after(0, self.combine_process_complete, success)
        except Exception as e:
            self.root.after(0, self.combine_process_error, str(e))
    
    def combine_process_complete(self, success):
        """Called when combine process is complete"""
        self.is_processing = False
        self.combine_button.config(state='normal')
        self.progress.stop()
        
        if success:
            self.status_var.set("Combination completed successfully!")
        else:
            self.status_var.set("Combination failed.")
    
    def combine_process_error(self, error_msg):
        """Called when combine process encounters an error"""
        self.is_processing = False
        self.combine_button.config(state='normal')
        self.progress.stop()
        self.status_var.set("Error occurred during processing.")
        self.log_message(f"Unexpected error: {error_msg}")
        messagebox.showerror("Error", f"Unexpected error: {error_msg}")
    
    def log_message(self, message):
        """Add a message to the log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        
        self.log_text.insert(tk.END, log_entry)
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def clear_log(self):
        """Clear the log text area"""
        self.log_text.delete(1.0, tk.END)
        self.log_message("Log cleared.")

def main():
    # Create the main window
    root = tk.Tk()
    
    # Set the application icon (if available)
    try:
        # Try to set an icon - this will work if you have an icon file
        if sys.platform.startswith('darwin'):  # macOS
            root.call('wm', 'iconbitmap', root._w, '-default', 'icon.icns')
        else:  # Windows
            root.iconbitmap('icon.ico')
    except:
        pass  # Icon not found, continue without it
    
    # Create the application
    app = ExcelCombinerGUI(root)
    
    # Handle window closing
    def on_closing():
        if app.is_processing:
            if messagebox.askokcancel("Quit", "Processing is in progress. Do you want to quit?"):
                root.destroy()
        else:
            root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    
    # Start the GUI event loop
    root.mainloop()

if __name__ == "__main__":
    main()