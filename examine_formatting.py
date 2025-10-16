#!/usr/bin/env python3
"""
Test script to examine Excel formatting in sample files
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def examine_excel_formatting(file_path):
    """Examine the formatting of an Excel file."""
    print(f"\n=== Examining {os.path.basename(file_path)} ===")
    
    try:
        # Load workbook with openpyxl to access formatting
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"Worksheet name: {ws.title}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
        
        # Check for filled cells and their colors
        filled_rows = []
        
        for row_idx in range(1, min(ws.max_row + 1, 20)):  # Check first 20 rows
            for col_idx in range(1, min(ws.max_column + 1, 5)):  # Check first 5 columns
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
                    if row_idx not in filled_rows:
                        filled_rows.append(row_idx)
                    
                    print(f"Row {row_idx}, Col {col_idx}: Fill pattern = {cell.fill.patternType}")
                    if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                        if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                            print(f"  Foreground color: {cell.fill.fgColor.rgb}")
                        if hasattr(cell.fill.fgColor, 'indexed') and cell.fill.fgColor.indexed:
                            print(f"  Indexed color: {cell.fill.fgColor.indexed}")
                    
                    # Get cell value
                    if cell.value:
                        print(f"  Cell value: {str(cell.value)[:50]}...")
        
        print(f"Found formatting in rows: {filled_rows}")
        
        # Also load with pandas to see the data
        df = pd.read_excel(file_path)
        print(f"DataFrame shape: {df.shape}")
        print("First few rows:")
        print(df.head(3))
        
    except Exception as e:
        print(f"Error examining {file_path}: {e}")

def main():
    sample_folder = "/Users/gr4yf1r3/Library/CloudStorage/OneDrive-Nuance/audioMover/_migration/walgreens_excelPlayground/ReDooV2/en_transcriptions_locationprompt_Tuned_9.15.2025_For_ScriptSplits_part1-4/sample_data"
    
    # Get all Excel files
    import glob
    excel_files = glob.glob(os.path.join(sample_folder, "*.xlsx"))
    excel_files.sort()
    
    for file_path in excel_files:
        examine_excel_formatting(file_path)

if __name__ == "__main__":
    main()