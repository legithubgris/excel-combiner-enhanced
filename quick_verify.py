#!/usr/bin/env python3
"""
Quick verification of GUI output formatting
"""

from openpyxl import load_workbook
import os

def verify_gui_output():
    file_path = "/Users/gr4yf1r3/Library/CloudStorage/OneDrive-Nuance/audioMover/_migration/walgreens_excelPlayground/ReDooV2/en_transcriptions_locationprompt_Tuned_9.15.2025_For_ScriptSplits_part1-4/sample_data/gui_test_combined.xlsx"
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    formatted_count = 0
    for row_idx in range(1, min(ws.max_row + 1, 100)):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
                formatted_count += 1
                if formatted_count <= 5:  # Show first 5
                    print(f"Row {row_idx}, Col {col_idx}: {cell.value} - Fill: {cell.fill.fgColor}")
    
    print(f"\nTotal formatted cells found: {formatted_count}")
    return formatted_count > 0

if __name__ == "__main__":
    success = verify_gui_output()
    print(f"\n{'✅ Formatting preserved!' if success else '❌ No formatting found'}")