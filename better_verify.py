#!/usr/bin/env python3
"""
Better verification focusing on actual highlight colors
"""

from openpyxl import load_workbook

def verify_highlight_colors(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    
    highlighted_cells = []
    
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if (cell.fill and 
                hasattr(cell.fill, 'fgColor') and 
                cell.fill.fgColor and 
                hasattr(cell.fill.fgColor, 'rgb') and 
                cell.fill.fgColor.rgb and 
                cell.fill.patternType != 'none' and
                cell.fill.fgColor.rgb not in ['00000000', 'FFFFFFFF']):  # Skip black/white defaults
                
                highlighted_cells.append({
                    'row': row_idx,
                    'col': col_idx, 
                    'color': cell.fill.fgColor.rgb,
                    'value': str(cell.value)[:30]
                })
    
    return highlighted_cells

# Check original file
original_highlights = verify_highlight_colors("/Users/gr4yf1r3/Library/CloudStorage/OneDrive-Nuance/audioMover/_migration/walgreens_excelPlayground/ReDooV2/en_transcriptions_locationprompt_Tuned_9.15.2025_For_ScriptSplits_part1-4/sample_data/en_transcriptions_locationprompt_Tuned_9.15.2025_For_ScriptSplits_part4.xlsx")

print("Original file highlights:")
for h in original_highlights:
    print(f"  Row {h['row']}, Col {h['col']}: {h['color']} - {h['value']}")

# Check combined file  
combined_highlights = verify_highlight_colors("/Users/gr4yf1r3/Library/CloudStorage/OneDrive-Nuance/audioMover/_migration/walgreens_excelPlayground/ReDooV2/en_transcriptions_locationprompt_Tuned_9.15.2025_For_ScriptSplits_part1-4/sample_data/gui_test_combined.xlsx")

print(f"\nCombined file highlights:")
for h in combined_highlights:
    print(f"  Row {h['row']}, Col {h['col']}: {h['color']} - {h['value']}")

print(f"\nOriginal: {len(original_highlights)} highlights")
print(f"Combined: {len(combined_highlights)} highlights")
print(f"Success: {len(combined_highlights) > 0}")