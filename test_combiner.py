#!/usr/bin/env python3
"""
Test script for Excel combiner with formatting preservation
"""

import os
import sys
import glob
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

def read_excel_data_with_formatting(file_path):
    """Read Excel file and return data with formatting information."""
    try:
        # Read the Excel file with pandas for data
        df = pd.read_excel(file_path)
        
        # Load with openpyxl to get formatting
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Get the filename without extension for the source column
        filename = os.path.basename(file_path)
        
        # Capture row formatting information
        row_formats = {}
        
        for row_idx in range(1, ws.max_row + 1):
            row_format = {}
            has_formatting = False
            
            # Check each cell in the row for formatting
            for col_idx in range(1, min(ws.max_column + 1, len(df.columns) + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                cell_format = {}
                
                # Check for fill (background color)
                if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
                    if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                        if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                            # Store the RGB value as string
                            rgb_val = cell.fill.fgColor.rgb
                            if hasattr(rgb_val, 'rgb'):
                                rgb_val = rgb_val.rgb
                            cell_format['fill_color'] = str(rgb_val)
                            has_formatting = True
                
                # Check for font formatting
                if cell.font:
                    font_info = {}
                    if cell.font.bold:
                        font_info['bold'] = True
                        has_formatting = True
                    if cell.font.italic:
                        font_info['italic'] = True
                        has_formatting = True
                    if cell.font.color and hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                        # Store the RGB value as string
                        rgb_val = cell.font.color.rgb
                        if hasattr(rgb_val, 'rgb'):
                            rgb_val = rgb_val.rgb
                        font_info['color'] = str(rgb_val)
                        has_formatting = True
                    
                    if font_info:
                        cell_format['font'] = font_info
                
                if cell_format:
                    row_format[col_idx] = cell_format
            
            if has_formatting:
                row_formats[row_idx] = row_format
        
        return df, filename, row_formats
        
    except Exception as e:
        print(f"Error reading file {file_path}: {str(e)}")
        return None, None, None

def combine_excel_files_with_formatting(folder_path, output_filename):
    """Combine multiple Excel files into one with preserved formatting."""
    
    # Get all Excel files in the folder
    excel_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
    excel_files.sort()
    
    if not excel_files:
        print(f"No Excel files found in folder: {folder_path}")
        return False
    
    print(f"Found {len(excel_files)} Excel files to combine:")
    for file in excel_files:
        print(f"  - {os.path.basename(file)}")
    
    try:
        # Create a new workbook for output
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.title = "Combined_Data"
        
        # Initialize variables for combining data
        combined_data = []
        all_formatting = []
        current_row = 1
        header_added = False
        
        for file_index, file_path in enumerate(excel_files):
            print(f"Processing: {os.path.basename(file_path)}")
            
            df, source_filename, row_formats = read_excel_data_with_formatting(file_path)
            
            if df is None:
                continue
                
            # Skip empty files
            if df.empty:
                print(f"  Skipping empty file: {source_filename}")
                continue
            
            # Ensure we have standard column names
            if len(df.columns) >= 3:
                # Use first 3 columns and add source column
                df_subset = df.iloc[:, :3].copy()
                df_subset.columns = ['Filename', 'Transcription', 'Status']
                df_subset['Source_File'] = ''
            else:
                print(f"  Warning: File {source_filename} has fewer than 3 columns. Skipping.")
                continue
            
            # Add source filename to the first data row of this file
            if len(df_subset) > 0:
                if not header_added:
                    # First file: include header
                    if len(df_subset) > 1:
                        df_subset.iloc[1, df_subset.columns.get_loc('Source_File')] = source_filename
                    header_added = True
                else:
                    # Subsequent files: skip header, add source to first data row
                    if len(df_subset) > 1:
                        df_subset = df_subset.iloc[1:].copy()  # Skip header row
                        if len(df_subset) > 0:
                            df_subset.iloc[0, df_subset.columns.get_loc('Source_File')] = source_filename
            
            # Store data and formatting info
            combined_data.append(df_subset)
            
            # Adjust row formatting indices for the combined file
            if row_formats and len(df_subset) > 0:
                adjusted_formats = {}
                for orig_row, format_info in row_formats.items():
                    # Calculate new row position in combined file
                    if file_index == 0:
                        # First file
                        new_row = current_row + (orig_row - 1)
                    else:
                        # Subsequent files: adjust for skipped header
                        if orig_row > 1:  # Skip header row
                            new_row = current_row + (orig_row - 2)
                        else:
                            continue  # Skip header formatting
                    
                    adjusted_formats[new_row] = format_info
                
                all_formatting.append(adjusted_formats)
                print(f"  Captured formatting for {len(adjusted_formats)} rows")
            
            current_row += len(df_subset)
            print(f"  Added {len(df_subset)} rows")
        
        if not combined_data:
            print("No data to combine!")
            return False
        
        # Combine all DataFrames
        final_df = pd.concat(combined_data, ignore_index=True)
        
        # Write data to the workbook
        for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                output_ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply formatting
        formatted_count = 0
        for format_dict in all_formatting:
            for row_num, row_format in format_dict.items():
                if row_num <= output_ws.max_row:
                    for col_num, cell_format in row_format.items():
                        if col_num <= output_ws.max_column:
                            cell = output_ws.cell(row=row_num, column=col_num)
                            
                            # Apply fill color
                            if 'fill_color' in cell_format:
                                try:
                                    fill_color = cell_format['fill_color']
                                    # Ensure it's a valid hex color
                                    if len(fill_color) == 8 and fill_color.startswith('FF'):
                                        fill_color = fill_color[2:]  # Remove alpha channel
                                    elif len(fill_color) != 6:
                                        continue  # Skip invalid colors
                                    
                                    cell.fill = PatternFill(start_color=fill_color, 
                                                          end_color=fill_color, 
                                                          fill_type='solid')
                                    formatted_count += 1
                                except Exception as e:
                                    print(f"    Warning: Could not apply fill color {fill_color}: {e}")
                            
                            # Apply font formatting (simplified - only bold/italic)
                            if 'font' in cell_format:
                                try:
                                    font_info = cell_format['font']
                                    current_font = cell.font
                                    
                                    new_font = Font(
                                        name=current_font.name or 'Calibri',
                                        size=current_font.size or 11,
                                        bold=font_info.get('bold', current_font.bold),
                                        italic=font_info.get('italic', current_font.italic)
                                        # Skip color for now due to complexity
                                    )
                                    cell.font = new_font
                                except Exception as e:
                                    print(f"    Warning: Could not apply font formatting: {e}")
        
        # Create output file path and save
        output_path = os.path.join(folder_path, output_filename)
        output_wb.save(output_path)
        
        print(f"Successfully combined {len(excel_files)} files with preserved formatting!")
        print(f"Output saved to: {output_path}")
        print(f"Total rows in combined file: {len(final_df)}")
        print(f"Columns: {list(final_df.columns)}")
        print(f"Applied formatting to {formatted_count} cells")
        
        return True
        
    except Exception as e:
        print(f"Error combining files: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    # Test with sample data
    sample_folder = "/Users/gr4yf1r3/Library/CloudStorage/OneDrive-Nuance/audioMover/_migration/walgreens_excelPlayground/ReDooV2/en_transcriptions_locationprompt_Tuned_9.15.2025_For_ScriptSplits_part1-4/sample_data"
    output_filename = "test_combined_with_formatting.xlsx"
    
    print("Testing Excel file combination with formatting preservation...")
    success = combine_excel_files_with_formatting(sample_folder, output_filename)
    
    if success:
        print("\n✅ Test completed successfully!")
        print(f"Check the output file: {os.path.join(sample_folder, output_filename)}")
    else:
        print("\n❌ Test failed!")

if __name__ == "__main__":
    main()