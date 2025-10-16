#!/usr/bin/env python3
"""
Clean test of the full-row highlighting with just sample data
"""

import sys
import os
sys.path.append('/Users/gr4yf1r3/Library/CloudStorage/OneDrive-Nuance/audioMover/_migration/walgreens_excelPlayground/ReDooV2/en_transcriptions_locationprompt_Tuned_9.15.2025_For_ScriptSplits_part1-4')

from excel_combiner_gui import ExcelCombinerGUI
import tkinter as tk

class CleanTestCombiner:
    def __init__(self):
        root = tk.Tk()
        root.withdraw()
        
        self.app = ExcelCombinerGUI(root)
        self.app.folder_path.set("/Users/gr4yf1r3/Library/CloudStorage/OneDrive-Nuance/audioMover/_migration/walgreens_excelPlayground/ReDooV2/en_transcriptions_locationprompt_Tuned_9.15.2025_For_ScriptSplits_part1-4/sample_data")
        self.app.output_filename.set("final_full_row_test.xlsx")
        
        # Override log_message to show progress
        original_log = self.app.log_message
        def console_log(message):
            if "Found formatting" in message or "Total rows" in message or "Applied full-row" in message:
                print(f"LOG: {message}")
            original_log(message)
        self.app.log_message = console_log
    
    def test_combine(self):
        print("Testing full-row highlighting with clean sample data...")
        
        # First, clean up any previous output files to ensure clean test
        sample_folder = "/Users/gr4yf1r3/Library/CloudStorage/OneDrive-Nuance/audioMover/_migration/walgreens_excelPlayground/ReDooV2/en_transcriptions_locationprompt_Tuned_9.15.2025_For_ScriptSplits_part1-4/sample_data"
        output_files_to_remove = ['gui_test_combined.xlsx', 'debug_combined.xlsx', 'test_combined_with_formatting.xlsx']
        
        for file in output_files_to_remove:
            file_path = os.path.join(sample_folder, file)
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"Cleaned up: {file}")
        
        result = self.app.combine_excel_files()
        return result

def main():
    tester = CleanTestCombiner()
    success = tester.test_combine()
    
    if success:
        print("\n✅ Clean test completed successfully!")
        
        # Verify the results
        from openpyxl import load_workbook
        output_file = "/Users/gr4yf1r3/Library/CloudStorage/OneDrive-Nuance/audioMover/_migration/walgreens_excelPlayground/ReDooV2/en_transcriptions_locationprompt_Tuned_9.15.2025_For_ScriptSplits_part1-4/sample_data/final_full_row_test.xlsx"
        
        if os.path.exists(output_file):
            wb = load_workbook(output_file)
            ws = wb.active
            
            print("\nFinal verification:")
            highlighted_rows = 0
            rows_with_wide_highlighting = 0
            
            for row_idx in range(1, min(ws.max_row + 1, 100)):
                colored_cols = 0
                for col_idx in range(1, 25):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if (cell.fill and cell.fill.patternType != 'none' and 
                        hasattr(cell.fill, 'fgColor') and cell.fill.fgColor and 
                        hasattr(cell.fill.fgColor, 'rgb') and 
                        cell.fill.fgColor.rgb not in ['00000000', 'FFFFFFFF']):
                        colored_cols += 1
                
                if colored_cols > 0:
                    highlighted_rows += 1
                    if colored_cols >= 15:
                        rows_with_wide_highlighting += 1
            
            print(f"📊 Results:")
            print(f"  - Total highlighted rows: {highlighted_rows}")
            print(f"  - Rows with full-width highlighting (15+ cols): {rows_with_wide_highlighting}")
            print(f"  - Total data rows: {ws.max_row - 1}")
            
            if rows_with_wide_highlighting > 0:
                print("🎉 Full-row highlighting is working perfectly!")
            else:
                print("❌ Full-row highlighting needs adjustment")
        
    else:
        print("\n❌ Clean test failed!")

if __name__ == "__main__":
    main()