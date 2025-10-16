#!/usr/bin/env python3
"""
Verify that formatting was preserved in the combined file
"""

from openpyxl import load_workbook
import os

def verify_formatting(file_path):
    """Verify that formatting was preserved in the combined file."""
    print(f"Verifying formatting in: {os.path.basename(file_path)}")
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    formatted_cells = []
    
    for row_idx in range(1, min(ws.max_row + 1, 50)):  # Check first 50 rows
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
                cell_value = str(cell.value)[:30] + "..." if cell.value and len(str(cell.value)) > 30 else str(cell.value)
                fill_color = "Unknown"
                
                if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                    if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
                        fill_color = cell.fill.fgColor.rgb
                
                formatted_cells.append({
                    'row': row_idx,
                    'col': col_idx,
                    'value': cell_value,
                    'color': fill_color
                })
    
    print(f"Found {len(formatted_cells)} formatted cells:")
    for cell in formatted_cells[:10]:  # Show first 10
        print(f"  Row {cell['row']}, Col {cell['col']}: '{cell['value']}' - Color: {cell['color']}")
    
    if len(formatted_cells) > 10:
        print(f"  ... and {len(formatted_cells) - 10} more")
    
    return len(formatted_cells) > 0

def main():
    output_file = "/Users/gr4yf1r3/Library/CloudStorage/OneDrive-Nuance/audioMover/_migration/walgreens_excelPlayground/ReDooV2/en_transcriptions_locationprompt_Tuned_9.15.2025_For_ScriptSplits_part1-4/sample_data/test_combined_with_formatting.xlsx"
    
    if os.path.exists(output_file):
        has_formatting = verify_formatting(output_file)
        if has_formatting:
            print("\n✅ Formatting verification successful!")
        else:
            print("\n❌ No formatting found in output file")
    else:
        print(f"❌ Output file not found: {output_file}")

if __name__ == "__main__":
    main()